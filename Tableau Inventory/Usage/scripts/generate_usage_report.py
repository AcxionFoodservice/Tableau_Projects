"""
Tableau Cloud Usage Report Generator
-------------------------------------
Produces an Excel workbook with three sheets:
  - Summary      : KPI cards, top-15 workbooks, per-project activity table
  - All Workbooks: one row per workbook with permissions and 90-day user data
  - By Project   : one row per top-level project with rollup stats

Usage:
    python generate_usage_report.py [--days N] [--output path/to/report.xlsx]

Environment variables required:
    TABLEAU_PAT_NAME    Personal Access Token name
    TABLEAU_PAT_SECRET  Personal Access Token secret
Optional:
    TABLEAU_SERVER      Default: https://us-east-1.online.tableau.com
    TABLEAU_SITE        Default: einstein
"""

import argparse
import os
import sys
import time
import zipfile
from collections import defaultdict
from datetime import date

import requests
import tableauhyperapi as ha
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add scripts dir to path so tableau_connect is importable when run directly
sys.path.insert(0, os.path.dirname(__file__))
from tableau_connect import (
    SERVER, API_VER, SITE_ID, NS,
    get_token, get_paged, get_one,
)

TS_EVENTS_DS_ID = "a2cfcec3-cfb9-426c-ae88-f27cb8675216"


# ── Styles ────────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MED_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
DARK_GRAY  = "595959"


def _font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", wrap=False):
    return Alignment(horizontal=h, vertical="center", wrap_text=wrap)


# ── Data fetching ─────────────────────────────────────────────────────────────

def fetch_90d_users(token: str) -> dict:
    """Download TS Events extract and return {(wb_name_lower, proj_name_lower): {count, users, wb_name, proj_name}}."""
    print("Downloading TS Events extract...", flush=True)
    r = requests.get(
        f"{SERVER}/api/{API_VER}/sites/{SITE_ID}/datasources/{TS_EVENTS_DS_ID}/content",
        headers={"x-tableau-auth": token},
        stream=True,
    )
    r.raise_for_status()
    tdsx_path = "/tmp/ts_events.tdsx"
    extract_dir = "/tmp/ts_events_extracted"
    with open(tdsx_path, "wb") as f:
        for chunk in r.iter_content(8192):
            f.write(chunk)

    with zipfile.ZipFile(tdsx_path) as z:
        hyper_name = next(n for n in z.namelist() if n.endswith(".hyper"))
        z.extractall(extract_dir)

    hyper_path = os.path.join(extract_dir, hyper_name)
    print("Querying TS Events for 90-day unique viewers...", flush=True)
    wb_90d = {}
    with ha.HyperProcess(telemetry=ha.Telemetry.DO_NOT_SEND_USAGE_DATA_TO_TABLEAU) as hyper:
        with ha.Connection(hyper.endpoint, hyper_path) as conn:
            rows = conn.execute_list_query("""
                SELECT DISTINCT workbook_name, item_project_name, actor_user_name
                FROM "public"."Extract"
                WHERE event_name IN ('Access View', 'Access Authoring View')
                  AND workbook_name IS NOT NULL
            """)
    raw = defaultdict(set)
    for wb_name, proj_name, user in rows:
        key = (str(wb_name).strip().lower(), str(proj_name or "").strip().lower())
        raw[key].add((str(wb_name), str(proj_name or ""), str(user)))

    for key, entries in raw.items():
        users = sorted({e[2] for e in entries})
        first = next(iter(entries))
        wb_90d[key] = {"count": len(users), "users": users,
                       "wb_name": first[0], "proj_name": first[1]}
    print(f"  90d data found for {len(wb_90d)} workbooks", flush=True)
    return wb_90d


def fetch_rest_data(token: str) -> tuple:
    """Return (projects, workbooks, wb_view_count, wb_perm_users, wb_perm_groups)."""
    print("Fetching users & groups...", flush=True)
    user_map = {}
    for u in get_paged(token, "users", "user"):
        name = u.get("fullName") or u.get("name", "?")
        email = u.get("name", "")
        user_map[u.get("id")] = f"{name} ({email})" if name != email else email

    group_map = {}
    for g in get_paged(token, "groups", "group"):
        group_map[g.get("id")] = g.get("name", "?")

    print("Fetching projects...", flush=True)
    projects = {}
    for p in get_paged(token, "projects", "project"):
        projects[p.get("id")] = {"name": p.get("name"), "parentId": p.get("parentProjectId")}

    def top_ancestor(pid):
        visited = set()
        while pid and projects.get(pid, {}).get("parentId"):
            pid = projects[pid]["parentId"]
            if pid in visited:
                break
            visited.add(pid)
        return pid

    proj_to_top = {pid: top_ancestor(pid) for pid in projects}

    print("Fetching workbooks...", flush=True)
    workbooks = {}
    for wb in get_paged(token, "workbooks", "workbook"):
        proj_el = wb.find(f"{{{NS}}}project")
        pid = proj_el.get("id") if proj_el is not None else None
        top_pid = proj_to_top.get(pid, pid) if pid else None
        workbooks[wb.get("id")] = {
            "name": wb.get("name"),
            "projectId": pid,
            "projectName": projects.get(pid, {}).get("name", "") if pid else "",
            "topProjectName": projects.get(top_pid, {}).get("name", "")
                              or projects.get(pid, {}).get("name", "(No Project)") if top_pid else "(No Project)",
        }
    print(f"  {len(workbooks)} workbooks", flush=True)

    print("Fetching view counts...", flush=True)
    wb_view_count = defaultdict(int)
    for v in get_paged(token, "views", "view", "&includeUsageStatistics=true"):
        wb_el = v.find(f"{{{NS}}}workbook")
        usage_el = v.find(f"{{{NS}}}usage")
        if wb_el is not None and usage_el is not None:
            wb_view_count[wb_el.get("id")] += int(usage_el.get("totalViewCount", 0))

    print(f"Fetching permissions for {len(workbooks)} workbooks...", flush=True)
    wb_perm_users, wb_perm_groups = {}, {}
    for i, wbid in enumerate(workbooks):
        if i % 25 == 0:
            print(f"  {i}/{len(workbooks)}...", flush=True)
        root = get_one(token, f"workbooks/{wbid}/permissions")
        users, groups = [], []
        for perm in root.findall(f".//{{{NS}}}granteeCapabilities"):
            user_el = perm.find(f"{{{NS}}}user")
            group_el = perm.find(f"{{{NS}}}group")
            caps = [c.get("name") for c in perm.findall(f".//{{{NS}}}capability") if c.get("mode") == "Allow"]
            if caps:
                if user_el is not None:
                    users.append(user_map.get(user_el.get("id", ""), user_el.get("id", "")))
                if group_el is not None:
                    groups.append(group_map.get(group_el.get("id", ""), group_el.get("id", "")))
        wb_perm_users[wbid] = sorted(set(users))
        wb_perm_groups[wbid] = sorted(set(groups))
        time.sleep(0.05)

    return projects, workbooks, wb_view_count, wb_perm_users, wb_perm_groups


# ── Row assembly ──────────────────────────────────────────────────────────────

def assemble_rows(projects, workbooks, wb_view_count, wb_perm_users, wb_perm_groups, wb_90d) -> list:
    rows = []
    for wbid, wb in workbooks.items():
        top_name = wb["topProjectName"]
        proj_name_lower = wb["projectName"].strip().lower()
        wb_name_lower = wb["name"].strip().lower()
        ts_key = (wb_name_lower, proj_name_lower)
        ts = wb_90d.get(ts_key)
        if ts is None:
            matches = [v for k, v in wb_90d.items() if k[0] == wb_name_lower]
            ts = matches[0] if len(matches) == 1 else None
        rows.append({
            "project": top_name,
            "workbook": wb["name"],
            "all_time_views": wb_view_count.get(wbid, 0),
            "users_90d": ts["count"] if ts else 0,
            "groups": ", ".join(wb_perm_groups.get(wbid, [])),
            "perm_users": ", ".join(wb_perm_users.get(wbid, [])),
            "users_90d_list": ", ".join(ts["users"]) if ts else "",
        })
    return rows


# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(rows: list, output_path: str, run_date: str):
    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    for col, width in zip("ABCDE", [3, 36, 18, 18, 3]):
        ws.column_dimensions[col].width = width

    ws.merge_cells("B1:D2")
    c = ws["B1"]
    c.value = "Tableau Cloud – Workbook Usage Report"
    c.font = _font(16, True, WHITE)
    c.fill = _fill(DARK_BLUE)
    c.alignment = _align("center")
    ws.row_dimensions[1].height = ws.row_dimensions[2].height = 22

    ws.merge_cells("B3:D3")
    c = ws["B3"]
    c.value = f"90-Day Window: TS Events  |  Generated: {run_date}"
    c.font = _font(10, False, WHITE)
    c.fill = _fill(MED_BLUE)
    c.alignment = _align("center")
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 10

    kpis = [
        ("Total Projects", len({r["project"] for r in rows})),
        ("Total Workbooks", len(rows)),
        ("Active (90d)",   sum(1 for r in rows if r["users_90d"] > 0)),
        ("Inactive (90d)", sum(1 for r in rows if r["users_90d"] == 0)),
    ]
    for (col, row_n), (label, val) in zip([("B",5),("D",5),("B",8),("D",8)], kpis):
        ws.merge_cells(f"{col}{row_n}:{chr(ord(col)+1)}{row_n+1}")
        c = ws[f"{col}{row_n}"]
        c.value = val; c.font = _font(22, True, DARK_BLUE)
        c.fill = _fill(LIGHT_BLUE); c.alignment = _align("center")
        ws.row_dimensions[row_n].height = 24; ws.row_dimensions[row_n+1].height = 14
        ws.merge_cells(f"{col}{row_n+2}:{chr(ord(col)+1)}{row_n+2}")
        c2 = ws[f"{col}{row_n+2}"]
        c2.value = label; c2.font = _font(9, False, DARK_GRAY)
        c2.fill = _fill(LIGHT_BLUE); c2.alignment = _align("center")
        ws.row_dimensions[row_n+2].height = 14
    ws.row_dimensions[11].height = 10

    ws["B12"].value = "Top 15 Workbooks by 90-Day Unique Users"
    ws["B12"].font = _font(12, True, DARK_BLUE)
    ws.row_dimensions[12].height = 18

    for ci, label in enumerate(["Workbook", "Project", "90d Users"], 2):
        c = ws.cell(row=13, column=ci)
        c.value = label; c.font = _font(10, True, WHITE)
        c.fill = _fill(MED_BLUE); c.alignment = _align("center", True); c.border = _border()
    ws.row_dimensions[13].height = 18

    for idx, r in enumerate(sorted(rows, key=lambda x: -x["users_90d"])[:15], 14):
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        for ci, val in enumerate([r["workbook"], r["project"], r["users_90d"]], 2):
            c = ws.cell(row=idx, column=ci)
            c.value = val; c.font = _font(10)
            c.fill = _fill(bg); c.border = _border()
            c.alignment = _align("center") if ci == 4 else _align("left", True)
        ws.row_dimensions[idx].height = 16

    ws.row_dimensions[29].height = 10
    ws["B30"].value = "Activity by Project"
    ws["B30"].font = _font(12, True, DARK_BLUE)
    ws.row_dimensions[30].height = 18

    for ci, label in enumerate(["Project", "Workbooks", "90d Active"], 2):
        c = ws.cell(row=31, column=ci)
        c.value = label; c.font = _font(10, True, WHITE)
        c.fill = _fill(MED_BLUE); c.alignment = _align("center", True); c.border = _border()
    ws.row_dimensions[31].height = 18

    proj_stats = defaultdict(lambda: {"total": 0, "active": 0})
    for r in rows:
        proj_stats[r["project"]]["total"] += 1
        if r["users_90d"] > 0:
            proj_stats[r["project"]]["active"] += 1

    for idx, (pname, pd) in enumerate(sorted(proj_stats.items()), 32):
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        for ci, val in enumerate([pname, pd["total"], pd["active"]], 2):
            c = ws.cell(row=idx, column=ci)
            c.value = val; c.font = _font(10)
            c.fill = _fill(bg); c.border = _border()
            c.alignment = _align("center") if ci in (3, 4) else _align("left", True)
        ws.row_dimensions[idx].height = 15

    # ── All Workbooks sheet ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("All Workbooks")
    ws2.sheet_view.showGridLines = False
    cols = [("Project",22),("Workbook",38),("All-Time Views",14),
            ("90d Unique Users",16),("Permission Groups",40),
            ("Permission Users",50),("90d Users",60)]
    for ci, (label, width) in enumerate(cols, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = width
        c = ws2.cell(row=1, column=ci)
        c.value = label; c.font = _font(10, True, WHITE)
        c.fill = _fill(MED_BLUE); c.alignment = _align("center", True); c.border = _border()
    ws2.row_dimensions[1].height = 20
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    for idx, r in enumerate(sorted(rows, key=lambda x: (x["project"], x["workbook"])), 2):
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        vals = [r["project"], r["workbook"], r["all_time_views"], r["users_90d"],
                r["groups"], r["perm_users"], r["users_90d_list"]]
        for ci, val in enumerate(vals, 1):
            c = ws2.cell(row=idx, column=ci)
            c.value = val; c.font = _font(10)
            c.fill = _fill(bg); c.border = _border()
            c.alignment = _align("center") if ci in (3, 4) else _align("left", True)
        if r["users_90d"] > 0:
            ws2.cell(row=idx, column=4).font = _font(10, True, MED_BLUE)
        ws2.row_dimensions[idx].height = 15

    # ── By Project sheet ───────────────────────────────────────────────────────
    ws3 = wb.create_sheet("By Project")
    ws3.sheet_view.showGridLines = False
    proj_cols = [("Project",28),("Total Workbooks",16),("Active 90d",14),
                 ("Inactive 90d",14),("% Active",12),
                 ("Total 90d User-Visits",22),("Top Workbook (90d)",38)]
    for ci, (label, width) in enumerate(proj_cols, 1):
        ws3.column_dimensions[get_column_letter(ci)].width = width
        c = ws3.cell(row=1, column=ci)
        c.value = label; c.font = _font(10, True, WHITE)
        c.fill = _fill(MED_BLUE); c.alignment = _align("center", True); c.border = _border()
    ws3.row_dimensions[1].height = 20
    ws3.freeze_panes = "A2"

    proj_detail = defaultdict(lambda: {"total":0,"active":0,"sum_90d":0,"top_wb":("",0)})
    for r in rows:
        p = r["project"]
        proj_detail[p]["total"] += 1
        proj_detail[p]["sum_90d"] += r["users_90d"]
        if r["users_90d"] > 0:
            proj_detail[p]["active"] += 1
        if r["users_90d"] > proj_detail[p]["top_wb"][1]:
            proj_detail[p]["top_wb"] = (r["workbook"], r["users_90d"])

    for idx, pname in enumerate(sorted(proj_detail), 2):
        pd = proj_detail[pname]
        bg = LIGHT_GRAY if idx % 2 == 0 else WHITE
        pct = f"{pd['active']/pd['total']*100:.0f}%" if pd["total"] else "0%"
        vals = [pname, pd["total"], pd["active"], pd["total"]-pd["active"],
                pct, pd["sum_90d"], pd["top_wb"][0]]
        for ci, val in enumerate(vals, 1):
            c = ws3.cell(row=idx, column=ci)
            c.value = val; c.font = _font(10)
            c.fill = _fill(bg); c.border = _border()
            c.alignment = _align("center") if ci in range(2,7) else _align("left", True)
        ws3.row_dimensions[idx].height = 15

    wb.save(output_path)
    print(f"Report saved: {output_path}", flush=True)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate Tableau Cloud usage report.")
    parser.add_argument("--output", default=f"tableau_usage_report_{date.today()}.xlsx",
                        help="Output .xlsx path (default: tableau_usage_report_YYYY-MM-DD.xlsx)")
    args = parser.parse_args()

    run_date = date.today().strftime("%B %d, %Y")

    print("Authenticating to Tableau Cloud...", flush=True)
    token = get_token()

    wb_90d = fetch_90d_users(token)
    projects, workbooks, wb_view_count, wb_perm_users, wb_perm_groups = fetch_rest_data(token)
    rows = assemble_rows(projects, workbooks, wb_view_count, wb_perm_users, wb_perm_groups, wb_90d)

    print(f"\nBuilding Excel report ({len(rows)} workbooks across {len({r['project'] for r in rows})} projects)...", flush=True)
    build_excel(rows, args.output, run_date)
    print("Done.", flush=True)


if __name__ == "__main__":
    main()
